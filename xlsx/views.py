from django.shortcuts import render
from django.views.generic.base import View
from .forms import XlsxModelForm
from .models import Xlsx
import openpyxl
from quiz.models import Quiz, Question, Answer
from django.contrib import messages


class UploadFileView(View):

    def get(self, request):
        form = XlsxModelForm(request.POST or None, request.FILES or None)
        return render(request, 'xlsx/upload.html', {'form': form})

    def post(self, request):
        form = XlsxModelForm(request.POST or None, request.FILES or None)
        if form.is_valid():
            form_obj = form.save(commit=False)
            form_obj.author = request.user
            form_obj.save()
            form_obj = XlsxModelForm()
            obj = Xlsx.objects.get(activated=False, author=request.user)
            try:
                book = openpyxl.open(obj.file_name.path, read_only=True)

                sheet = book.worksheets[0]
                for row in range(2, sheet.max_row+1):
                    if Quiz.objects.filter(name=sheet[row][0].value,
                                           desc=sheet[row][1].value,
                                           number_of_questions=int(sheet[row][2].value),
                                           time=int(sheet[row][3].value)).exists():
                        obj.activated = True
                        obj.save()
                        messages.add_message(request, messages.ERROR, 'The object already exists!')
                        return render(request, 'xlsx/upload.html', {'form': form})

                    Quiz.objects.create(
                        name=sheet[row][0].value,
                        desc=sheet[row][1].value,
                        number_of_questions=int(sheet[row][2].value),
                        time=int(sheet[row][3].value)
                    )

                sheet_1 = book.worksheets[1]
                for row in range(2, sheet_1.max_row + 1):
                    quiz_obj = Quiz.objects.get(name=sheet_1[row][1].value)
                    Question.objects.create(
                        content=sheet_1[row][0].value,
                        quiz=quiz_obj
                    )

                sheet_2 = book.worksheets[2]
                for row in range(2, sheet_2.max_row + 1):
                    question_obj = Question.objects.get(content=sheet_2[row][2].value)
                    Answer.objects.create(
                        content=sheet_2[row][0].value,
                        correct=bool(sheet_2[row][1].value),
                        question=question_obj
                    )
                messages.add_message(request, messages.INFO, 'Adding successfully')
            except Exception:
                messages.add_message(request, messages.ERROR, 'Incorrectly filled file!')
            obj.activated = True
            obj.save()
        return render(request, 'xlsx/upload.html', {'form': form_obj})

