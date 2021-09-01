import openpyxl
from quiz.models import Quiz, Question, Answer


class AddQuizMixin:

    def add_quiz(self, xlsx):
        try:
            book = openpyxl.open(xlsx.file_name.path, read_only=True)

            sheet = book.worksheets[0]
            for row in range(2, sheet.max_row+1):
                if Quiz.objects.filter(name=sheet[row][0].value,
                                       desc=sheet[row][1].value,
                                       number_of_questions=int(sheet[row][2].value),
                                       time=int(sheet[row][3].value)).exists():
                    xlsx.delete()
                    return False

                Quiz.objects.create(
                    name=sheet[row][0].value,
                    desc=sheet[row][1].value,
                    number_of_questions=int(sheet[row][2].value),
                    time=int(sheet[row][3].value)
                )

            sheet_1 = book.worksheets[1]
            for row in range(2, sheet_1.max_row + 1):
                quiz_obj = Quiz.objects.get(name=sheet_1[row][1].value)
                Question.objects.create(
                    content=sheet_1[row][0].value,
                    quiz=quiz_obj
                )

            sheet_2 = book.worksheets[2]
            for row in range(2, sheet_2.max_row + 1):
                question_obj = Question.objects.get(content=sheet_2[row][2].value)
                Answer.objects.create(
                    content=sheet_2[row][0].value,
                    correct=bool(sheet_2[row][1].value),
                    question=question_obj
                )
            xlsx.activated = True
            xlsx.save()
            return True
        except Exception:
            xlsx.delete()
            return False
